import os
import csv
import sys
import json
import argparse
import difflib
from argparse import ArgumentParser
import traceback
from collections import namedtuple
import openpyxl
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.cell import Cell, MergedCell


MetadataIds = namedtuple("MetadataIds", "indicators, countries, combos")

# DE Names
SEL_MONTHLY_NAME = 'Mean monthly subsistence expenditure line (cost of meeting basic needs)'
CTP_MONTHLY_NAME = 'Mean monthly capacity to pay for health care'
SHARE_HH_WITH_OOP_TOTAL_NAME = 'Share of households with out-of-pocket payments for health care (total)'
SHARE_HH_NO_OOP_TOTAL_NAME = 'Share of households without out-of-pocket payments for health care (total)'
SHARE_HH_WITH_OOP_QUINTILE_NAME = 'Share of households with out-of-pocket payments for health care (by consumption quintile)'
SHARE_HH_NO_OOP_QUINTILE_NAME = 'Share of households without out-of-pocket payments for health care (by consumption quintile)'
GGHED_CHE_NAME = 'Public spending on health as a share of current spending on health'
VHI_CHE_NAME = 'Voluntary health insurance spending as a share of current spending on health'
OOP_CHE_NAME = 'Out-of-pocket payments as a share of current spending on health (oop)'
OTHER_CHE_NAME = 'Other spending as a share of current spending on health'
POVERTY_LINE_OLD_NAME = 'Percent below subsistence expenditure line'
CATA_HEALTHCARE_TOTAL_NAME = 'Breakdown of catastrophic health spending by type of health care (total)'
CATA_QUINTILE_NAME = 'Share of households with catastrophic health spending by consumption quintile'
CATA_TOTAL_NAME = 'Share of households with catastrophic health spending (total)'
FURTHERIMPOV_CATA_NAME = 'Share of households with catastrophic health spending who are further impoverished'
IMPOV_CATA_NAME = 'Share of households with catastrophic health spending who are impoverished'
GGHED_GGE_NAME = 'Public spending on health as a share of government spending'
UN_EUSILC_DENTAL_QUINTILE_NAME = 'Self-reported unmet need for dental care due to cost, distance and waiting time (quintile)'

COUNTRY_DICT = {
    'BIH': 'Bosnia and Herzegovina',
    'CZH': 'Czech Republic',
    'DEU': 'Federal Republic of Germany',
    'FRA': 'French Republic',
    'GEO': 'Georgia',
    'LUX': 'Grand Duchy of Luxembourg',
    'GRC': 'Hellenic Republic',
    'HUN': 'Hungary',
    'IRL': 'Ireland',
    'BEL': 'Kingdom of Belgium',
    'DNK': 'Kingdom of Denmark',
    'NOR': 'Kingdom of Norway',
    'SPA': 'Kingdom of Spain',
    'SWE': 'Kingdom of Sweden',
    'NET': 'Kingdom of the Netherlands',
    'KGZ': 'Kyrgyz Republic',
    'MNE': 'Montenegro',
    'POR': 'Portuguese Republic',
    'AND': 'Principality of Andorra',
    'MCO': 'Principality of Monaco',
    'ALB': 'Republic of Albania',
    'ARM': 'Republic of Armenia',
    'AUT': 'Republic of Austria',
    'AZE': 'Republic of Azerbaijan',
    'BLR': 'Republic of Belarus',
    'BUL': 'Republic of Bulgaria',
    'CRO': 'Republic of Croatia',
    'CYP': 'Republic of Cyprus',
    'EST': 'Republic of Estonia',
    'FIN': 'Republic of Finland',
    'ICE': 'Republic of Iceland',
    'ITA': 'Republic of Italy',
    'KAZ': 'Republic of Kazakhstan',
    'LVA': 'Republic of Latvia',
    'LTU': 'Republic of Lithuania',
    'MAT': 'Republic of Malta',
    'MDA': 'Republic of Moldova',
    'MKD': 'Republic of North Macedonia',
    'POL': 'Republic of Poland',
    'SMR': 'Republic of San Marino',
    'SRB': 'Republic of Serbia',
    'SVN': 'Republic of Slovenia',
    'TJK': 'Republic of Tajikistan',
    'TUR': 'Republic of Türkiye',
    'UZB': 'Republic of Uzbekistan',
    'ROU': 'Romania',
    'RUS': 'Russian Federation',
    'SVK': 'Slovak Republic',
    'ISR': 'State of Israel',
    'SWI': 'Swiss Confederation',
    'TKM': 'Turkmenistan',
    'UKR': 'Ukraine',
    'UNK': 'United Kingdom of Great Britain and Northern Ireland',
}


COMBO_LIST = [
    'Total, Outpatient care',
    'Poorest, Dental care',
    'Medical products, Richest',
    '2nd',
    '3rd',
    'Poorest, Outpatient care',
    'Dental care, 4th',
    'Diagnostic tests, 3rd',
    'Diagnostic tests, Poorest',
    'Medical products, 3rd',
    'Medical products, Poorest',
    '2nd, Outpatient care',
    'Medicines, 3rd',
    'Inpatient care',
    'Outpatient care, 3rd',
    'default',
    'Medical products, Total',
    'Dental care',
    'Richest, Dental care',
    'Richest, Outpatient care',
    'Richest',
    'Inpatient care, 4th',
    'Dental care, 2nd',
    'Diagnostic tests, Richest',
    'Medicines, 4th',
    '2nd, Medicines',
    'Inpatient care, 3rd',
    'Total, Medicines',
    'Inpatient care, Total',
    'Dental care, Total',
    'Medical products, 2nd',
    'Medical products',
    'Poorest',
    'Diagnostic tests, 4th',
    'Outpatient care',
    'Dental care, 3rd',
    'Total',
    'Richest, Inpatient care',
    'Medicines',
    'Medical products, 4th',
    'Diagnostic tests',
    'Diagnostic tests, 2nd',
    'Inpatient care, 2nd',
    '4th',
    'Richest, Medicines',
    'Poorest, Medicines',
    'Diagnostic tests, Total',
    'Poorest, Inpatient care',
    'Outpatient care, 4th'
]


OLD_NAMES_DICT = {
    'Mean annual per capita OOP by structure (by quintile)': {
        'Medicines': 'Annual out-of-pocket payments for outpatient medicines per person by consumption quintile',
        'Inpatient care': 'Annual out-of-pocket payments for inpatient care per person by consumption quintile',
        'Dental care': 'Annual out-of-pocket payments for dental care person by consumption quintile',
        'Outpatient care': 'Annual out-of-pocket payments for outpatient care per person by consumption quintile',
        'Diagnostic tests': 'Annual out-of-pocket payments for diagnostic tests per person by consumption quintile',
        'Medical products': 'Annual out-of-pocket payments for medical products per person by consumption quintile',
    },
    'Catastrophic out-of-pocket payments (total)': {
        'NA': CATA_TOTAL_NAME,
        'Medicines': CATA_HEALTHCARE_TOTAL_NAME,
        'Inpatient care': CATA_HEALTHCARE_TOTAL_NAME,
        'Dental care': CATA_HEALTHCARE_TOTAL_NAME,
        'Outpatient care': CATA_HEALTHCARE_TOTAL_NAME,
        'Diagnostic tests': CATA_HEALTHCARE_TOTAL_NAME,
        'Medical products': CATA_HEALTHCARE_TOTAL_NAME,
    },
    'Catastrophic out-of-pocket payments (by qunitile)': {
        'NA': CATA_QUINTILE_NAME,
        'Medicines': 'Breakdown of catastrophic health spending by type of health care (by consumption quintile)',
        'Inpatient care': 'Breakdown of catastrophic health spending by type of health care (by consumption quintile)',
        'Dental care': 'Breakdown of catastrophic health spending by type of health care (by consumption quintile)',
        'Outpatient care': 'Breakdown of catastrophic health spending by type of health care (by consumption quintile)',
        'Diagnostic tests': 'Breakdown of catastrophic health spending by type of health care (by consumption quintile)',
        'Medical products': 'Breakdown of catastrophic health spending by type of health care (by consumption quintile)',
    },
    'Mean annual per capita OOP by structure (total)': 'Annual out-of-pocket payments on health care per person by type of health care (total)',
    'At risk of impoverishment (all households)': 'Share of households at risk of impoverishment (all households)',
    'further impoverished (all households)': 'Share of further impoverished households (total)',
    'Impoverished (all households)': 'Share of Impoverished households (all households)',
    'Impoverishing health spending': 'Share of households with impoverishing health spending',
    'At risk of impoverishment (catastrophic households)': 'Share of households with catastrophic health spending who at risk of impoverishment',
    'further impoverished (catastrophic households)': FURTHERIMPOV_CATA_NAME,
    'Impoverished (catastrophic households)': IMPOV_CATA_NAME,
    'Not at risk of impoverishment (catastrophic households)': 'Share of households with catastrophic health spending who are not at risk of impoverishment',
    'Out-of-pocket payments as a share of total household spending among households with catastrophic spending (by quintile)': 'Out-of-pocket payments as a share of total household spending among households with catastrophic health spending by consumption quintile',
    'Average out-of-pocket payments as a share of total household spending among further impoverished households': 'Out-of-pocket payments as a share of total household spending among further impoverished households',
    'Mean annual capacity to pay': CTP_MONTHLY_NAME,
    POVERTY_LINE_OLD_NAME: 'Percent below subsistence expenditure line (basic needs line)',
    'Mean annual subsistence expenditure line': SEL_MONTHLY_NAME,
    'Share of households without out-of-pocket payments (by quintile)': SHARE_HH_NO_OOP_QUINTILE_NAME,
    'Share of households without out-of-pocket payments (total)': SHARE_HH_NO_OOP_TOTAL_NAME,
    'Share of households with out-of-pocket payments (by quintile)': SHARE_HH_WITH_OOP_QUINTILE_NAME,
    'Share of households with out-of-pocket payments (total)': SHARE_HH_WITH_OOP_TOTAL_NAME,
    'Mean annual per capita OOP (by quintile)': 'Annual out-of-pocket payments for health care per person (by consumption quintile)',
    'Mean annual per capita OOP (total)': 'Annual out-of-pocket payments for health care per person (total)',
    'Out-of-pocket payments for health care as a share of household consumption (by quintile)': 'Out-of-pocket payments for health care as a share of household consumption (by consumption quintile)',
    'Share of total OOP by structure (total population)': 'Breakdown of out-of-pocket payments by type of health care (total)',
    'Share of OOP by structure (by quintile)': 'Breakdown of out-of-pocket payments by type of health care (by consumption quintile)',
}


INDICATOR_IGNORING_SERVICE = [
    'Annual out-of-pocket payments for outpatient medicines per person by consumption quintile',
    'Annual out-of-pocket payments for inpatient care per person by consumption quintile',
    'Annual out-of-pocket payments for dental care person by consumption quintile',
    'Annual out-of-pocket payments for outpatient care per person by consumption quintile',
    'Annual out-of-pocket payments for diagnostic tests per person by consumption quintile',
    'Annual out-of-pocket payments for medical products per person by consumption quintile',
]

INDICATOR_IGNORING_QUINTILE = [
    CATA_HEALTHCARE_TOTAL_NAME,
    'Public spending on health as a share of current spending on health by type of care',
    'Annual out-of-pocket payments on health care per person by type of health care (total)',
    'Out-of-pocket payments as a share of current spending on health by type of care',
    'Breakdown of out-of-pocket payments by type of health care (total)',
    'Voluntary health insurance spending as a share of current spending on health by type of care',
]

COC_DEFAULT_ID = ""
COC_TOTAL_ID = ""


def get_new_name(indicator_name: str, service: str):
    """Maps old data elements names to new ones

    Args:
        indicator_name (str): Old data element name
        service (str): data element categoryOptionCombos name

    Raises:
        ValueError: If a DE can't be mapped 

    Returns:
        new_indicator_name (str): New data element name
    """

    if indicator_name in OLD_NAMES_DICT:
        debug(f'Found old name: {indicator_name}')
        new_name = OLD_NAMES_DICT[indicator_name]
        if isinstance(new_name, dict):
            if service not in new_name:
                raise ValueError('get_new_name: cant map old indicator')

        return new_name[service] if isinstance(new_name, dict) else new_name
    else:
        return indicator_name


def make_combo_string(quintile: str, service: str):
    """Maps quintile and service to a categoryOptionCombos name

    Args:
        quintile (str): categoryOptions name
        service (str): categoryOptionCombos name

    Returns:
        combo (str): categoryOptionCombos name
    """

    if service == 'NA':
        result = quintile
        if quintile == "Total":
            result = "default"
    elif quintile == 'NA':
        result = service
    else:
        combo = quintile + ', ' + service
        combo_alt = service + ', ' + quintile

        if combo in COMBO_LIST:
            result = combo
        elif combo_alt in COMBO_LIST:
            result = combo_alt

    return result


def check_for_empty_csv_fields(**elements):
    """Checks if there are empty values in the CSV file and prints warning
    """

    for name, value in elements.items():
        if name != 'row' and not value:
            print(f'WARNING: Empty {name} variable in CSV file in row:\n{elements["row"]}')


def currency_converter(amount: str, country: str, year: str, figure: str):
    """Applies currency conversion to the CSV file values

    Args:
        amount (str): Original value
        country (str): Country code
        year (str): Year string
        figure (str): Figure code

    Returns:
        adjusted_amount (str): New value with currency conversion applied
    """

    currency_figures = ["F5", "F9", "F10a", "F10b", "F10c", "F10d", "F10e", "F10f", "F26"]

    debug(f'currency_converter amount: {amount} | country_code: {country} | year: {year} | figure: {figure}')

    if figure not in currency_figures:
        debug('currency_converter figure not in list')
        return amount

    # NOTE: Temporal fixes until the CURRENCY_TABLE gets fixed
    if country == "GRC":
        country = "GRE"
    elif country == "DNK":
        country = "DEN"
    elif country == "IRL":
        country = "IRE"
    elif country == "ROU":
        country = "ROM"

    try:
        coefficient = CURRENCY_TABLE[
            (CURRENCY_TABLE['code'] == country)
        ][year].values[0]
    except KeyError:
        # If no coefficient available for year, get the closest year to present
        last_year = next(reversed(CURRENCY_TABLE.keys()))
        coefficient = CURRENCY_TABLE[
            (CURRENCY_TABLE['code'] == country)
        ][last_year].values[0]
    except Exception:
        traceback.print_exc()
        sys.exit(1)

    debug('currency_converter coefficient:', coefficient)

    adjusting_for_inflation = round(float(amount) / coefficient, 2)

    return str(adjusting_for_inflation)


def get_csv_indicator_value(value: str, real_value: str):
    """Checks if real_value exists and its not "NA" if --real_value flag is set

    Args:
        value (str): CSV Value field
        real_value (str): CSV real_value field

    Returns:
        real_value (str): Appropriate value based on --real_value flag
    """

    if REAL_VALUE:
        return real_value if real_value != "NA" else value

    return value


def create_dict_if_dont_exist(dictionary: dict, key: str):
    """check if key is in nested dictionary and creates a new empty dict if its not

    Args:
        dict (dict): Nested dictionary
        key (str): dictionary to check

    Returns:
        dictionary: Updated nested dictionary
    """

    if key not in dictionary:
        dictionary[key] = {}
    return dictionary


def extract_values_from_csv(filename: str):
    """Given a CSV file name creates a dictionary of the CSV file data

    Args:
        filename (str): CSV file name

    Returns:
        values (dict): Dictionary with the CSV file data
    """

    values = {}

    try:
        with open(filename, 'r', encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                indicator_name = row['indicator_name']
                country = row['country']
                year = row['year']
                quintile = row['quintile']
                service = row['service']
                if CURRENCY and indicator_name != POVERTY_LINE_OLD_NAME:
                    figure = row['figure_id']
                    value = currency_converter(row['value'], country, year, figure)
                else:
                    value = get_csv_indicator_value(row['value'], row['real_value'])

                check_for_empty_csv_fields(
                    row=row,
                    indicator_name=indicator_name,
                    country=country,
                    year=year,
                    quintile=quintile,
                    service=service,
                    value=value
                )

                indicator_name = get_new_name(indicator_name, service)

                country_name = COUNTRY_DICT[country]

                values = create_dict_if_dont_exist(values, country_name)
                values[country_name] = create_dict_if_dont_exist(values[country_name], year)
                values[country_name][year] = create_dict_if_dont_exist(values[country_name][year], indicator_name)

                service = 'NA' if indicator_name in INDICATOR_IGNORING_SERVICE else service
                quintile = 'NA' if indicator_name in INDICATOR_IGNORING_QUINTILE else quintile

                cat_opt_combo = make_combo_string(quintile, service)
                if value != 'NA' and value is not None:
                    values[country_name][year][indicator_name][cat_opt_combo] = value
                else:
                    debug('Empty CSV value in row: ', row)

        return values
    except Exception:
        traceback.print_exc()
        sys.exit(1)


def get_metadata_ids(workbook: Workbook):
    """Crates a named tuple containing dictionaries with the ids of indicators, 
    countries and combos used, the ids are extracted from the bulk load template

    Args:
        workbook (Workbook): XLSX file with the bulk load template

    Returns:
        ids (MetadataIds): named tuple containing dictionaries with the ids of indicators, countries and combos used
    """

    global COC_DEFAULT_ID, COC_TOTAL_ID

    indicators_id_dict = {}
    countries_id_dict = {}
    combos_id_dict = {}
    sheet = workbook['Metadata']

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
        identifier = row[0]
        type_col = row[1]
        name = str(row[2]).strip()

        if type_col == 'categoryOptionCombos':
            combos_id_dict[identifier] = name

            if name == "default":
                COC_DEFAULT_ID = identifier
            if name == "Total":
                COC_TOTAL_ID = identifier

        if type_col == 'dataElements':
            indicators_id_dict[name] = identifier

        if type_col == 'organisationUnit':
            countries_id_dict[name] = identifier

    return MetadataIds(indicators_id_dict, countries_id_dict, combos_id_dict)


def get_indicator_id(ids: MetadataIds, name: str):
    """Gets the ID of the provided DE and raises an error if the ID is not found

    Args:
        ids (MetadataIds): named tuple containing dictionaries with the ids of indicators, countries and combos used
        name (str): Data element name

    Returns:
        id (str): ID of the provided DE
    """
    try:
        return ids.indicators[name]
    except KeyError:
        print(f'ERROR: Data element "{name}" can\'t be matched with an ID, check metadata')
        print(f'Closest candidates: {difflib.get_close_matches(name, ids.indicators.keys())}')
        return None


def update_latest_dict(latest_pre_2019_des_dict: dict, indicator_name: str, year: str, combo_id: str, value: str):
    """Check if the DE need to populate a latest version of it and returns the updated latest_pre_2019_des_dict

    Args:
        latest_pre_2019_des_dict (dict): Dictionary storing the latest year, combo and value
        indicator_name (str): Data element name
        year (str): Data element value year
        combo_id (str): Data element value combo
        value (str): Data element value
    """

    latest_year = None
    if indicator_name in latest_pre_2019_des_dict:
        latest_year = latest_pre_2019_des_dict[indicator_name][0]

        if int(latest_year) < int(year):
            latest_pre_2019_des_dict[indicator_name] = [year, combo_id, value]


def store_latest_data(data: dict, ids: MetadataIds, latest_pre_2019_des_dict: dict, country_id: str):
    """Stores the data values for the latest DEs for the given country

    Args:
        data (dict): Dictionary with the matched data
        ids (MetadataIds): named tuple containing dictionaries with the ids of indicators, countries and combos used
        latest_pre_2019_des_dict (dict): Dictionary storing the latest year, combo and value
        country_id (str): Data country ID
    """

    for indicator_name, latest_list in latest_pre_2019_des_dict.items():
        latest_year, latest_combo, latest_value = latest_list
        if latest_year != "0":
            latest_indicator_name = indicator_name + " - 2019 or LAY"
            last_indicator_id = get_indicator_id(ids, latest_indicator_name)
            if not last_indicator_id:
                continue

            data[country_id][latest_year] = create_dict_if_dont_exist(data[country_id][latest_year], last_indicator_id)

            data[country_id][latest_year][last_indicator_id][latest_combo] = latest_value
            debug(
                f'pre_2019_de_names check: "{latest_indicator_name}" | {latest_year}" | {data[country_id][latest_year][last_indicator_id][latest_combo]}'
            )


def make_matched_values(csv_values_dict: dict, ids: MetadataIds):
    """Maps the formNames and data of csv_values_dict with the metadata ids, applies direct monthly transformation

    Args:
        csv_values_dict (dict): Dictionary with the CSV data
        ids (MetadataIds): named tuple containing dictionaries with the ids of indicators, countries and combos used

    Returns:
        (dict): Dictionary with the CSV values indexed by metadata ids
    """

    data = {}

    for country, country_data in csv_values_dict.items():
        latest_pre_2019_des_dict = {
            CATA_HEALTHCARE_TOTAL_NAME: ["0", "", ""],
            OOP_CHE_NAME: ["0", "", ""],
            GGHED_GGE_NAME: ["0", "", ""],
            CATA_QUINTILE_NAME: ["0", "", ""],
            CATA_TOTAL_NAME: ["0", "", ""],
            FURTHERIMPOV_CATA_NAME: ["0", "", ""],
            IMPOV_CATA_NAME: ["0", "", ""],
            UN_EUSILC_DENTAL_QUINTILE_NAME: ["0", "", ""]
        }

        country_id = ids.countries[country]
        data[country_id] = {}

        for year, indicators in country_data.items():
            data[country_id] = create_dict_if_dont_exist(data[country_id], year)

            for indicator_name, indicator_combos in indicators.items():
                indicator_id = get_indicator_id(ids, indicator_name)
                if not indicator_id:
                    continue

                data[country_id][year] = create_dict_if_dont_exist(data[country_id][year], indicator_id)

                store_transformation_de(indicator_name, indicator_id)

                for combo_name, value in indicator_combos.items():
                    combo_ids = []
                    for id_code, name in ids.combos.items():
                        if name == combo_name:
                            combo_ids.append(id_code)
                    combo_id = '|'.join(combo_ids)

                    if check_mean_monthly_indicator(indicator_name):
                        debug("check_mean_monthly_indicator: ", indicator_name, value, float(value)/12)
                        value = str(float(value)/12)

                    update_latest_dict(latest_pre_2019_des_dict, indicator_name, year, combo_id, value)

                    data[country_id][year][indicator_id][combo_id] = value

        store_latest_data(data, ids, latest_pre_2019_des_dict, country_id)

    debug("pre_2019_de_names: ", dump_json_var(latest_pre_2019_des_dict))
    return data


# TRANSFORMATIONS
SHARE_HH_WITH_OOP_TOTAL = None
SHARE_HH_NO_OOP_TOTAL = None
SHARE_HH_WITH_OOP_QUINTILE = None
SHARE_HH_NO_OOP_QUINTILE = None
GGHED_CHE = None
VHI_CHE = None
OOP_CHE = None
OTHER_CHE = None


def check_mean_monthly_indicator(indicator_name: str):
    """Check if data element needs the monthly transformation

    Args:
        indicator_name (str): Form Name of the data element

    Returns:
        (bool): Boolean value of the check
    """

    mean_monthly_names = [SEL_MONTHLY_NAME, CTP_MONTHLY_NAME]

    return indicator_name in mean_monthly_names


def store_transformation_de(indicator_name: str, indicator_id: str):
    """Stores the data elements ids needed for transformations 

    Args:
        indicator_name (str): data element form name
        indicator_id (str): data element id
    """

    global SHARE_HH_WITH_OOP_TOTAL, SHARE_HH_NO_OOP_TOTAL, SHARE_HH_WITH_OOP_QUINTILE, SHARE_HH_NO_OOP_QUINTILE
    global GGHED_CHE, VHI_CHE, OOP_CHE, OTHER_CHE

    if indicator_name == SHARE_HH_WITH_OOP_TOTAL_NAME:
        SHARE_HH_WITH_OOP_TOTAL = indicator_id
    elif indicator_name == SHARE_HH_NO_OOP_TOTAL_NAME:
        SHARE_HH_NO_OOP_TOTAL = indicator_id
    elif indicator_name == SHARE_HH_WITH_OOP_QUINTILE_NAME:
        SHARE_HH_WITH_OOP_QUINTILE = indicator_id
    elif indicator_name == SHARE_HH_NO_OOP_QUINTILE_NAME:
        SHARE_HH_NO_OOP_QUINTILE = indicator_id
    elif indicator_name == GGHED_CHE_NAME:
        GGHED_CHE = indicator_id
    elif indicator_name == VHI_CHE_NAME:
        VHI_CHE = indicator_id
    elif indicator_name == OOP_CHE_NAME:
        OOP_CHE = indicator_id
    elif indicator_name == OTHER_CHE_NAME:
        OTHER_CHE = indicator_id


def get_indicator_value(matched_values: dict, country_id: str, year: str, indicator_id: str, combo_id: str, default: str | None = None):
    """_summary_

    Args:
        matched_values (dict): Dictionary with the values indexed by metadata id
        country_id (str): Value country ID
        year (str): Value year
        indicator_id (str): Value data element ID
        combo_id (str): Value combo ID
        default (str | None, optional): Value to return if not found, print find error if not provided. Defaults to None.

    Returns:
        value (str): Found value or default
    """

    try:
        return matched_values[country_id][year][indicator_id][combo_id]
    except KeyError:
        if default:
            return default

        print(f'ERROR: Can\'t find value for country: {country_id} year: {year} de: {indicator_id} combo: {combo_id}')
        return None


def get_spending_share_indicator(matched_values: dict, ids: dict, de: str, name: str):
    """Tries to get the data element value and prints a warning if no value can be found

    Args:
        matched_values (dict): Dictionary with the values indexed by metadata id
        ids (dict): Dictionary with the requested data element country, year and combo
        de (str): Data element id
        name (str): Data element name

    Returns:
        (str | None): Value of the requested data element or None in case of error
    """

    try:
        return float(matched_values[ids["country_id"]][ids["year"]][de][ids["combo_id"]])
    except KeyError:
        print(f'WARNING: Data element "{name}" for OU {ids["country_id"]} - {ids["year"]} is missing')
        return None


def make_transformations(matched_values: dict):
    """Performs transformations for the 'Share of households with out-of-pocket payments for health care' and 
    'Other spending as a share of current spending on health'

    Args:
        matched_values (dict): Dictionary with the values indexed by metadata id
    """

    households_ids = {
        SHARE_HH_WITH_OOP_TOTAL: SHARE_HH_NO_OOP_TOTAL,
        SHARE_HH_WITH_OOP_QUINTILE: SHARE_HH_NO_OOP_QUINTILE
    }

    for country_id, country_data in matched_values.items():
        for year, indicators in country_data.items():
            for indicator_id, indicator_combos in indicators.items():
                if indicator_id in households_ids.keys():
                    debug(f"make_transformations: {country_id} - {year} - {indicator_id}")

                    without_id = households_ids[indicator_id]
                    if not bool(indicator_combos):
                        indicator_combos = matched_values[country_id][year][without_id]

                    for combo_id, _ in indicator_combos.items():
                        without_value = get_indicator_value(matched_values, country_id, year, without_id, combo_id)

                        if without_value:
                            create_dict_if_dont_exist(matched_values[country_id][year][indicator_id], combo_id)
                            matched_values[country_id][year][indicator_id][combo_id] = str(100 - float(without_value))
                            debug(
                                f"make_transformations calc: {country_id} - {year} - {indicator_id} - {matched_values[country_id][year][indicator_id][combo_id]}"
                            )

                        gghed_che_value = get_spending_share_indicator(matched_values, ids, GGHED_CHE, GGHED_CHE_NAME)
                        vhi_che_value = get_spending_share_indicator(matched_values, ids, VHI_CHE, VHI_CHE_NAME)
                        oop_che_value = get_spending_share_indicator(matched_values, ids, OOP_CHE, OOP_CHE_NAME)

                        if gghed_che_value and vhi_che_value and oop_che_value:
                            matched_values[country_id][year][indicator_id][combo_id] = str(
                                100-(gghed_che_value + vhi_che_value + oop_che_value)
                            )
                        else:
                            print(
                                f'WARNING: Data element "{OTHER_CHE_NAME}" for OU {country_id} - {year} is missing values for transformation'
                            )
                            matched_values[country_id][year][indicator_id][combo_id] = ""


def write_org_unit(last_cell: Cell, matched_values: dict):
    """Writes the countries in the CSV data to the bulk load file

    Args:
        last_cell (Cell): Previous cell of the column
        matched_values (dict): Dictionary with the values indexed by metadata id
    """

    for country_id, country_data in matched_values.items():
        for _ in country_data:
            new_cell = last_cell.offset(row=1, column=0)
            new_cell.value = f'=_{country_id}'

            last_cell = new_cell


def write_years(last_cell: Cell, matched_values: dict):
    """Writes the years in the CSV data to the bulk load file

    Args:
        last_cell (Cell): Previous cell of the column
        matched_values (dict): Dictionary with the values indexed by metadata id
    """

    for _, country_data in matched_values.items():
        for year in country_data:
            new_cell = last_cell.offset(row=1, column=0)
            new_cell.value = year

            last_cell = new_cell


def write_indicator(col_indicator: str, col_combo: str, last_cell: Cell, matched_values: dict):
    """Writes the data elements in the CSV data to the bulk load file

    Args:
        col_indicator (str): Id of the data element
        col_combo (str): Id of the data elements combo
        last_cell (Cell): Previous cell of the column
        matched_values (dict): Dictionary with the values indexed by metadata id

    Returns:
        (int): Number of data elements added to the bulk load file
    """

    count = 0

    for _, country_data in matched_values.items():
        for __, indicators in country_data.items():
            for indicator_id, indicator_combos in indicators.items():
                if indicator_id == col_indicator:
                    for combo_id, value in indicator_combos.items():
                        ids = combo_id.split('|') if '|' in combo_id else combo_id
                        if col_combo in ids or (col_combo == COC_DEFAULT_ID and combo_id == COC_TOTAL_ID):
                            new_cell = last_cell.offset(row=1, column=0)
                            new_cell.value = value

                            last_cell = new_cell

                            count += 1

    return count


def write_values(workbook: Workbook, matched_values: dict):
    """Writes the CSV data to a new bulk load file using workbook as a template

    Args:
        workbook (Workbook): XLSX file with the bulk load template
        matched_values (dict): Dictionary with the values indexed by metadata id

    Returns:
        (int): Number of data elements added to the bulk load file
    """

    sheet = workbook['Data Entry']
    workbook.active = workbook['Data Entry']
    count = 0

    for index, col in enumerate(sheet.iter_cols(min_row=4)):
        if index == 0:
            last_cell = col[-1]
            write_org_unit(last_cell, matched_values)
        if index == 1:
            last_cell = col[-1]
            write_years(last_cell, matched_values)
        if index == 2:
            pass
        if index > 2:
            if not isinstance(col[0], MergedCell):
                col_indicator = str(col[0].value).rsplit('=_', maxsplit=1)[-1]
            col_combo = str(col[1].value).rsplit('=_', maxsplit=1)[-1]
            last_cell = col[-1]

            count += write_indicator(col_indicator, col_combo,
                                     last_cell, matched_values)

    workbook.save(OUT_FILENAME)

    debug(f'excel count: {count}')
    return count


def debug(*msg):
    """Writes the debug message to LOG_FILE
    """

    if DEBUG:
        with open(LOG_FILE, "a", encoding="utf-8") as log_file:
            print(*msg, file=log_file)


def dump_json_var(var: any):
    """Transforms var object to a JSON string

    Args:
        var (any): Object to be transformed

    Returns:
        (str): JSON string of the object
    """

    return json.dumps(var, indent=2)


def get_matched_values_len(matched_values: dict):
    """Gets the number of matched values from the CSV

    Args:
        matched_values (dict): Dictionary with the values indexed by metadata id

    Returns:
        (int): Number of matched values
    """

    lenght = 0

    for years in matched_values.values():
        for indicators in years.values():
            for combos in indicators.values():
                lenght += len(combos)
    return lenght


def filepath_exists(filepath: str):
    """Checks if path exists and its a file

    Args:
        filepath (str): Path to the file

    Returns:
        (bool): Value of the check
    """

    return os.path.isfile(filepath)


def get_template_path(parser: ArgumentParser, xlsx_template: str):
    """Returns a path to the bulk load template, if the --xlsx_template is not used returns the defaut template path.
    If the path is not valid prints error and exits.  

    Args:
        parser (ArgumentParser): Script argument parser. 
        xlsx_template (str): Value of --xlsx_template argument

    Returns:
        (str): Path to the template
    """

    if not xlsx_template:
        if filepath_exists(DEFAULT_TEMPLATE):
            xlsx_template = DEFAULT_TEMPLATE
        else:
            parser.error(f'The default template: {DEFAULT_TEMPLATE} doesn\'t exist')
    elif not filepath_exists(xlsx_template):
        parser.error(f'The template: {xlsx_template} doesn\'t exist')

    return xlsx_template


OUT_FILENAME = ''
DEFAULT_TEMPLATE = 'Quantitative_Data_UHCPW_Template.xlsx'
REAL_VALUE = False
DEBUG = False
LOG_FILE = 'log.json'
CURRENCY = False
CURRENCY_TABLE = None


def main():
    parser = argparse.ArgumentParser(description='Process CSV from "Data Extraction Tool" into "Bulk Load" XLSX files. \
                                     The script needs a template, it either can be supplied with the --xlsx_template \
                                     argument or by placing a template named "Quantitative_Data_UHCPW_Template.xlsx" \
                                     in the same folder as the script.\
                                     Outputs to a XLSX file with same name as the source CSV one.')
    parser.add_argument('indicators_csv', type=str, help='Source CSV file')
    parser.add_argument('-x', '--xlsx_template', type=str,
                        help='Bulk Load Quantitative XLSX template file path, if empty tries to open "Quantitative_Data_UHCPW_Template.xlsx"')
    adjusted_values_args = parser.add_mutually_exclusive_group()
    adjusted_values_args.add_argument('-r', '--real_value', action='store_true',
                                      help='Use real_value (if not NA) instead of value from the CSV source file, cant be used with -c/--currency')
    adjusted_values_args.add_argument('-c', '--currency', action='store_true',
                                      help='Apply currency adjustment to the applicable values, cant be used with -r/--real_value')
    parser.add_argument('-d', '--debug', action='store_true',
                        help='Display debug logs, its recommended to redirect the output into a file, e.g: ... > log.txt')
    args = parser.parse_args()

    if not filepath_exists(args.indicators_csv):
        parser.error(f'The source file: {args.indicators_csv} doesn\'t exist')

    global OUT_FILENAME, REAL_VALUE, DEBUG, CURRENCY, CURRENCY_TABLE
    OUT_FILENAME = f'{args.indicators_csv.split(".csv")[0]}.xlsx'
    REAL_VALUE = args.real_value
    DEBUG = args.debug
    CURRENCY = args.currency

    if DEBUG:
        f = open(LOG_FILE, 'w', encoding="utf-8")
        f.close()

    args.xlsx_template = get_template_path(parser, args.xlsx_template)

    debug('Source file:', args.indicators_csv)
    debug('Template:', args.xlsx_template)
    debug('Output file:', OUT_FILENAME)

    if CURRENCY:
        CURRENCY_TABLE = pd.read_csv(
            "https://docs.google.com/spreadsheets/d/1lEHQ9i-LO7gl0RWaJgYcfOJHjPefVXJbhgJ0Gn3iUPQ/export?format=csv&gid=56805701",
            decimal=","
        )
        debug('Currency table:', CURRENCY_TABLE)

    try:
        wb = openpyxl.load_workbook(args.xlsx_template)
    except Exception as e:
        debug("openpyxl exception: ", e)
        traceback.print_exc()
        sys.exit(1)

    csv_values_dict = extract_values_from_csv(args.indicators_csv)
    debug('csv_values_dict:\n ', dump_json_var(csv_values_dict))

    ids = get_metadata_ids(wb)

    debug(f'indicators ids:\n len: {len(ids.indicators)}\n values:\n', dump_json_var(ids.indicators))
    debug(f'countries ids:\n len: {len(ids.countries)}\n values:\n', dump_json_var(ids.countries))
    debug(f'combos ids:\n len: {len(ids.combos)}\n values:\n', dump_json_var(ids.combos))

    matched_values = make_matched_values(csv_values_dict, ids)

    csv_count = get_matched_values_len(matched_values)
    debug(f'matched_values count: {csv_count}\n')
    debug('matched_values:\n', dump_json_var(matched_values))

    make_transformations(matched_values)

    excel_count = write_values(wb, matched_values)
    debug(f'write_values count: {excel_count}\n')

    print(f'Processed {csv_count} entries from CSV file, written {excel_count} values to EXCEL')


if __name__ == '__main__':
    main()
